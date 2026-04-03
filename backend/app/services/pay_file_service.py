"""Monthly pay file generation service."""

import os
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.models.transaction import DealTransaction
from app.models.pay_file import PayFile
from app.config import settings


HEADER_FILL = PatternFill(start_color="1D1D1D", end_color="1D1D1D", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D7D0"),
)


def generate_pay_file(db: Session, period: str, adjustments: list[dict] | None = None) -> PayFile:
    """Generate a monthly pay file for the given period (e.g. '2026-03')."""
    # Get all pending transactions for the period
    year, month = period.split("-")
    txns = db.query(DealTransaction).filter(
        DealTransaction.payment_status == "pending",
        func.strftime("%Y-%m", DealTransaction.retail_date) == period
    ).all()

    if not txns:
        # Also try without date filter for pending transactions
        txns = db.query(DealTransaction).filter(
            DealTransaction.payment_status == "pending"
        ).all()

    # Apply adjustments
    exclusions = set()
    amount_overrides = {}
    if adjustments:
        for adj in adjustments:
            if adj.get("exclude"):
                exclusions.add(adj["transaction_id"])
            if adj.get("new_amount") is not None:
                amount_overrides[adj["transaction_id"]] = Decimal(str(adj["new_amount"]))

    # Filter and compute
    included_txns = [t for t in txns if t.id not in exclusions]
    total_amount = Decimal("0")
    for t in included_txns:
        amt = amount_overrides.get(t.id, t.support_amount or Decimal("0"))
        total_amount += amt

    # Create PayFile record
    pf = PayFile(
        period=period,
        total_amount=total_amount,
        total_units=len(included_txns),
        status="draft",
    )
    db.add(pf)
    db.flush()

    # Generate Excel
    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, included_txns, amount_overrides)

    # Sheet 2: Detail
    ws_detail = wb.create_sheet("Detail")
    _write_detail_sheet(ws_detail, included_txns, amount_overrides)

    # Sheet 3: Adjustments
    if adjustments:
        ws_adj = wb.create_sheet("Adjustments")
        _write_adjustments_sheet(ws_adj, adjustments)

    # Save file
    os.makedirs(settings.OUTPUT_DIR, exist_ok=True)
    filename = f"PayFile_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(settings.OUTPUT_DIR, filename)
    wb.save(filepath)

    pf.file_path = filepath

    # Link transactions
    for t in included_txns:
        t.pay_file_id = pf.id

    db.commit()
    return pf


def _style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _write_summary_sheet(ws, transactions, overrides):
    """Dealer-level summary."""
    headers = ["Ship-To Code", "Dealer Name", "Region", "Units", "Total Support"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    # Group by dealer
    dealer_totals = {}
    for t in transactions:
        key = t.dealer_ship_to or "UNKNOWN"
        if key not in dealer_totals:
            dealer_totals[key] = {"name": t.dealer_name, "units": 0, "total": Decimal("0")}
        dealer_totals[key]["units"] += 1
        amt = overrides.get(t.id, t.support_amount or Decimal("0"))
        dealer_totals[key]["total"] += amt

    row = 2
    for code, data in sorted(dealer_totals.items()):
        ws.cell(row=row, column=1, value=code).font = CELL_FONT
        ws.cell(row=row, column=2, value=data["name"]).font = CELL_FONT
        ws.cell(row=row, column=3, value="").font = CELL_FONT
        ws.cell(row=row, column=4, value=data["units"]).font = CELL_FONT
        ws.cell(row=row, column=5, value=float(data["total"])).font = CELL_FONT
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 20


def _write_detail_sheet(ws, transactions, overrides):
    """VIN-level detail."""
    headers = ["Ship-To Code", "Dealer Name", "VIN", "Campaign Code", "Material",
               "Retail Date", "Support Amount", "Sales Order", "Channel"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for row, t in enumerate(transactions, 2):
        amt = overrides.get(t.id, t.support_amount or Decimal("0"))
        values = [
            t.dealer_ship_to, t.dealer_name, t.vin, t.campaign_code,
            t.material, str(t.retail_date) if t.retail_date else "",
            float(amt), t.sales_order, t.channel,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
        ws.cell(row=row, column=7).number_format = '$#,##0.00'

    for i, w in enumerate([15, 30, 20, 12, 15, 12, 15, 15, 10], 1):
        ws.column_dimensions[chr(64 + i)].width = w


def _write_adjustments_sheet(ws, adjustments):
    """Manual adjustments log."""
    headers = ["Transaction ID", "New Amount", "Excluded", "Reason"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for row, adj in enumerate(adjustments, 2):
        ws.cell(row=row, column=1, value=adj.get("transaction_id", "")).font = CELL_FONT
        ws.cell(row=row, column=2, value=adj.get("new_amount")).font = CELL_FONT
        ws.cell(row=row, column=3, value="Yes" if adj.get("exclude") else "No").font = CELL_FONT
        ws.cell(row=row, column=4, value=adj.get("reason", "")).font = CELL_FONT
