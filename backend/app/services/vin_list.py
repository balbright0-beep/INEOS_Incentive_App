"""Excel parser for vin_specific program VIN lists.

Input shape: a workbook with a single sheet that has a header row plus
two columns — VIN and an amount column. The amount column header is
flexible (e.g. "MSRP REBATE", "Amount", "Rebate") so admins don't have
to rename their source spreadsheet before uploading.
"""

from io import BytesIO
from typing import Iterable
from openpyxl import load_workbook


VIN_LEN = 17

# Header keywords that identify the amount column. Match is case-
# insensitive and substring-based, so "MSRP REBATE", "Rebate Amount",
# "Total Amount", "Per-VIN Amount" all resolve to the same column.
AMOUNT_HEADER_HINTS = ("rebate", "amount", "incentive", "msrp", "discount")
VIN_HEADER_HINTS = ("vin",)


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _find_columns(header_row: Iterable) -> tuple[int | None, int | None]:
    """Find the (vin_col_idx, amount_col_idx) in a header row. Returns
    None for either index when no match. Indices are 0-based positions
    within the row tuple."""
    vin_idx = None
    amount_idx = None
    for i, cell in enumerate(header_row):
        h = _normalize_header(cell)
        if not h:
            continue
        if vin_idx is None and any(k in h for k in VIN_HEADER_HINTS):
            vin_idx = i
            continue
        if amount_idx is None and any(k in h for k in AMOUNT_HEADER_HINTS):
            amount_idx = i
    return vin_idx, amount_idx


def parse_vin_list(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Parse an Excel workbook of (VIN, amount) rows.

    Returns (rows, warnings) where rows is a list of {vin, amount} and
    warnings collects per-row diagnostics (bad VIN length, missing
    amount, etc.) so the upload UI can surface them inline. The full
    file is parsed even when some rows are bad — the admin gets the
    rejected list back instead of a single fail-everything error."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    if not wb.sheetnames:
        return [], ["Workbook has no sheets"]
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], ["Workbook is empty"]

    vin_idx, amount_idx = _find_columns(header)
    if vin_idx is None:
        # Fall back to position 0 — common for a 2-column file with a
        # plain "VIN" header that doesn't match any hint above.
        vin_idx = 0
    if amount_idx is None:
        # Fall back to the first column that ISN'T the VIN column. With
        # a strict 2-column file (the MSRP REBATE.xlsx shape) this
        # always picks position 1.
        amount_idx = 1 if vin_idx != 1 else 0

    out: list[dict] = []
    warnings: list[str] = []
    seen: dict[str, float] = {}
    for line_num, row in enumerate(rows_iter, start=2):
        if row is None:
            continue
        # Skip blank rows (all cells None or empty string).
        if all((c is None or (isinstance(c, str) and not c.strip())) for c in row):
            continue
        try:
            vin_raw = row[vin_idx]
            amt_raw = row[amount_idx]
        except IndexError:
            warnings.append(f"Row {line_num}: missing column")
            continue

        vin = str(vin_raw).strip().upper() if vin_raw is not None else ""
        if not vin:
            warnings.append(f"Row {line_num}: missing VIN")
            continue
        if len(vin) != VIN_LEN:
            warnings.append(f"Row {line_num}: VIN '{vin}' is {len(vin)} chars (expected {VIN_LEN})")
            continue

        if amt_raw is None or (isinstance(amt_raw, str) and not amt_raw.strip()):
            warnings.append(f"Row {line_num}: VIN {vin} has no amount")
            continue
        try:
            amount = float(amt_raw)
        except (TypeError, ValueError):
            warnings.append(f"Row {line_num}: VIN {vin} amount '{amt_raw}' is not a number")
            continue
        if amount <= 0:
            warnings.append(f"Row {line_num}: VIN {vin} amount {amount} is not positive")
            continue

        if vin in seen:
            # Duplicate VIN in the upload — last value wins, but warn
            # so the admin knows the source spreadsheet has dupes.
            warnings.append(
                f"Row {line_num}: VIN {vin} appears more than once "
                f"(prior amount {seen[vin]}, kept latest {amount})"
            )
        seen[vin] = amount

    out = [{"vin": v, "amount": a} for v, a in seen.items()]
    return out, warnings
