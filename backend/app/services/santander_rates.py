"""Parse Santander APR and Lease input files for rate lookups.

Source priority:
  1. SantanderRateFile DB row (uploaded via Settings page) — current
     month's rates without a redeploy.
  2. Filesystem fallback — repo-bundled copy. Lets the system keep
     serving rates if the DB row is missing or there's no upload yet
     (e.g., new install before the first admin upload).

Parsed rates are cached in process memory keyed by (kind, source-key).
The Settings upload endpoint calls invalidate_cache() so a new upload
takes effect immediately — no app restart needed.
"""

import io
import os
import glob
from openpyxl import load_workbook


# CPOS code to trim name mapping
CPOS_TO_TRIM = {
    "BASE": "Base",
    "FIELD": "Fieldmaster",
    "BLACK": "Belstaff",
    "TRIAL": "Trialmaster",
    "HIGHL": "Highlands",
}


# In-process parsed-row cache keyed by (kind, source_key). source_key
# changes when a new upload lands (uploaded_at timestamp) or when the
# filesystem copy is swapped (file mtime), so a fresh parse fires
# automatically on either change.
_PARSED_CACHE: dict[tuple[str, str], list[dict]] = {}


def invalidate_cache() -> None:
    """Drop the parsed-rate cache. Called by the upload endpoint so
    newly-uploaded rates take effect on the next lookup."""
    _PARSED_CACHE.clear()


def _load_xlsx_bytes(kind: str, base_dir: str) -> tuple[bytes | None, str]:
    """Return (xlsx_bytes, source_key) for a given kind. Tries the DB
    upload first, falls back to the filesystem copy. source_key is
    used as the cache key so we re-parse when either source changes.

    kind: 'apr' | 'lease' | 'state_apr' | 'state_lease'
    """
    # 1. DB upload
    try:
        from app.database import SessionLocal
        from app.models.rate_file import SantanderRateFile
        with SessionLocal() as db:
            row = db.query(SantanderRateFile).filter(SantanderRateFile.kind == kind).first()
            if row and row.data:
                # uploaded_at changes on every replace -> cache invalidates naturally
                key = f"db:{row.uploaded_at.isoformat() if row.uploaded_at else 'na'}"
                return bytes(row.data), key
    except Exception:
        # DB not ready (e.g. during startup before tables exist) — fall through.
        pass

    # 2. Filesystem fallback
    patterns = {
        "apr": "INEOS_APRInput_*.xlsx",
        "lease": "INEOS_LeaseInput_*.xlsx",
        "state_apr": "State_INEOS_APRInput_*.xlsx",
        "state_lease": "State_INEOS_LeaseInput_*.xlsx",
    }
    pattern = patterns.get(kind)
    if not pattern:
        return None, ""
    matches = glob.glob(os.path.join(base_dir, pattern))
    if not matches:
        return None, ""
    path = max(matches, key=os.path.getmtime)
    try:
        with open(path, "rb") as f:
            data = f.read()
        key = f"fs:{path}:{int(os.path.getmtime(path))}"
        return data, key
    except OSError:
        return None, ""


def _find_input_files(base_dir: str) -> dict:
    """Legacy filesystem-only lookup retained for any caller that still
    references it directly. New code should use _load_xlsx_bytes which
    also checks the DB upload."""
    files = {"apr": None, "lease": None, "lease_state": None}
    patterns = {
        "apr": "INEOS_APRInput_*.xlsx",
        "lease": "INEOS_LeaseInput_*.xlsx",
        "lease_state": "State_INEOS_LeaseInput_*.xlsx",
    }
    for key, pattern in patterns.items():
        matches = glob.glob(os.path.join(base_dir, pattern))
        if matches:
            files[key] = max(matches, key=os.path.getmtime)
    return files


def _model_code_to_params(model_code: str) -> dict:
    body = "station_wagon" if model_code.startswith("G01") else "quartermaster"
    suffix = model_code[-1] if model_code else "C"
    my = "MY26" if suffix == "D" else "MY25"
    return {"body_style": body, "model_year": my}


def _to_date_str(v) -> str | None:
    """Normalize a Start/End Date cell to YYYY-MM-DD. Inputs come back
    as datetime.datetime when openpyxl can parse them; some files
    occasionally store the date as a string."""
    if v is None:
        return None
    try:
        return v.strftime("%Y-%m-%d")  # datetime.datetime
    except AttributeError:
        s = str(v).strip()
        return s or None


def _parse_apr_xlsx(data: bytes) -> list[dict]:
    """Parse APR xlsx bytes into rate records. Pulled out so the same
    parser handles both the National (apr) and State (state_apr) files
    — same column layout, different region values."""
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb["Retail"] if "Retail" in wb.sheetnames else wb[wb.sheetnames[0]]
    rates: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        region = row[1]
        tier = row[2]
        term = row[3]
        model_code = str(row[9] or "")
        cpos = str(row[10] or "").upper()
        start_date = _to_date_str(row[13])  # Start Date
        end_date = _to_date_str(row[14])    # End Date
        rate = row[15]

        if rate is None or str(rate).strip().lower() == "std.":
            continue
        try:
            rate_val = float(rate)
        except (ValueError, TypeError):
            continue

        params = _model_code_to_params(model_code)
        trim = CPOS_TO_TRIM.get(cpos, cpos.title() if cpos else "Base")

        rates.append({
            "region": str(region) if region is not None else None,
            "tier": int(tier) if tier else 1,
            "term": int(term) if term else 0,
            "model_code": model_code,
            "body_style": params["body_style"],
            "model_year": params["model_year"],
            "trim": trim,
            "apr": rate_val,
            "start_date": start_date,
            "end_date": end_date,
        })
    wb.close()
    return rates


def _parse_lease_xlsx(data: bytes) -> list[dict]:
    """Parse Lease xlsx bytes into rate records. Same parser for
    National (lease) and State (state_lease) — both have the
    Region column."""
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb["Lease"] if "Lease" in wb.sheetnames else wb[wb.sheetnames[0]]
    rates: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        region = row[1]
        tier = row[2]
        term = row[3]
        model_code = str(row[8] or "")
        cpos = str(row[9] or "").upper()
        start_date = _to_date_str(row[12])  # Start Date
        end_date = _to_date_str(row[13])    # End Date
        mf = row[14]
        acq_fee = row[15]
        residual = row[19]  # Published INEOS Residual

        mf_val = None
        if mf is not None and str(mf).strip().lower() != "std.":
            try:
                mf_val = float(mf)
            except (ValueError, TypeError):
                pass

        res_val = None
        if residual is not None:
            try:
                res_val = float(str(residual).replace("%", ""))
            except (ValueError, TypeError):
                pass

        params = _model_code_to_params(model_code)
        trim = CPOS_TO_TRIM.get(cpos, cpos.title() if cpos else "Base")

        rates.append({
            "region": str(region) if region is not None else None,
            "tier": int(tier) if tier else 1,
            "term": int(term) if term else 0,
            "model_code": model_code,
            "body_style": params["body_style"],
            "model_year": params["model_year"],
            "trim": trim,
            "money_factor": mf_val,
            "money_factor_display": str(mf) if mf else "Std.",
            "residual_pct": res_val,
            "acq_fee": float(acq_fee) if acq_fee else 895,
            "start_date": start_date,
            "end_date": end_date,
        })
    wb.close()
    return rates


def _cached_rates(kind: str, base_dir: str, parser) -> list[dict]:
    data, source_key = _load_xlsx_bytes(kind, base_dir)
    if not data:
        return []
    cache_key = (kind, source_key)
    cached = _PARSED_CACHE.get(cache_key)
    if cached is not None:
        return cached
    parsed = parser(data)
    _PARSED_CACHE[cache_key] = parsed
    return parsed


def load_apr_rates(base_dir: str) -> list[dict]:
    """Return parsed APR rates. National + state files merged; callers
    that want state-only can filter on the 'region' field."""
    national = _cached_rates("apr", base_dir, _parse_apr_xlsx)
    state = _cached_rates("state_apr", base_dir, _parse_apr_xlsx)
    return national + state


def load_lease_rates(base_dir: str) -> list[dict]:
    """Return parsed lease rates. National + state files merged."""
    national = _cached_rates("lease", base_dir, _parse_lease_xlsx)
    state = _cached_rates("state_lease", base_dir, _parse_lease_xlsx)
    return national + state


# Region values used by the input files. Numeric '101' = national
# rate sheet; everything else is a 2-letter state code.
NATIONAL_REGION = "101"


def _filter_for_state(rows: list[dict], state: str | None) -> list[dict]:
    """Return only rows that apply to the customer's state. State rows
    win when present for that state — and we still keep national rows
    so the per-term picker can fall back when the state file is silent
    on a particular (term, tier) combo."""
    if not state:
        return [r for r in rows if r.get("region") == NATIONAL_REGION]
    state_upper = state.upper()
    return [r for r in rows if r.get("region") in (NATIONAL_REGION, state_upper)]


def _is_state_row(r: dict, state: str | None) -> bool:
    return bool(state and r.get("region") == state.upper())


def _pick_per_term_apr(matching: list[dict], state: str | None) -> dict[int, dict]:
    """For each term, pick a state row over a national row. Among rows
    with the same region, prefer the lowest APR — same tiebreak as
    before, just within-region instead of across all regions."""
    by_term: dict[int, dict] = {}
    for r in matching:
        t = r["term"]
        cur = by_term.get(t)
        if cur is None:
            by_term[t] = r
            continue
        new_is_state = _is_state_row(r, state)
        cur_is_state = _is_state_row(cur, state)
        if new_is_state and not cur_is_state:
            by_term[t] = r
        elif new_is_state == cur_is_state and r["apr"] < cur["apr"]:
            by_term[t] = r
    return by_term


def _pick_per_term_lease(matching: list[dict], state: str | None) -> dict[int, dict]:
    """Same shape as _pick_per_term_apr but tiebreak on money_factor
    (lower MF = better deal). None money_factors lose to numeric ones."""
    def mf_better(new: dict, cur: dict) -> bool:
        if new["money_factor"] is None:
            return False
        if cur["money_factor"] is None:
            return True
        return new["money_factor"] < cur["money_factor"]

    by_term: dict[int, dict] = {}
    for r in matching:
        t = r["term"]
        cur = by_term.get(t)
        if cur is None:
            by_term[t] = r
            continue
        new_is_state = _is_state_row(r, state)
        cur_is_state = _is_state_row(cur, state)
        if new_is_state and not cur_is_state:
            by_term[t] = r
        elif new_is_state == cur_is_state and mf_better(r, cur):
            by_term[t] = r
    return by_term


def get_apr_for_config(base_dir: str, model_year: str, body_style: str,
                       tier: int = 1, trim: str = None, state: str = None) -> list[dict]:
    """Get APR rates. If state is given, prefer state-specific rows
    over national for each (term, tier); otherwise return national only.
    Deduplicated by term."""
    all_rates = _filter_for_state(load_apr_rates(base_dir), state)
    matching = [
        r for r in all_rates
        if r["model_year"] == model_year
        and r["body_style"] == body_style
        and r["tier"] == tier
    ]
    if trim:
        trim_match = [r for r in matching if r["trim"].lower() == trim.lower()]
        if trim_match:
            matching = trim_match
    return sorted(_pick_per_term_apr(matching, state).values(), key=lambda r: r["term"])


def get_lease_for_config(base_dir: str, model_year: str, body_style: str,
                         tier: int = 1, trim: str = None, state: str = None) -> list[dict]:
    """Get lease rates. State preference + per-term dedup, matching
    get_apr_for_config."""
    all_rates = _filter_for_state(load_lease_rates(base_dir), state)
    matching = [
        r for r in all_rates
        if r["model_year"] == model_year
        and r["body_style"] == body_style
        and r["tier"] == tier
    ]
    if trim:
        trim_match = [r for r in matching if r["trim"].lower() == trim.lower()]
        if trim_match:
            matching = trim_match
    return sorted(_pick_per_term_lease(matching, state).values(), key=lambda r: r["term"])


def get_all_lease_by_trim(base_dir: str, model_year: str, body_style: str,
                          tier: int = 1, state: str = None) -> dict[str, list[dict]]:
    """Get lease rates grouped by trim for comparison display.
    State-aware in the same way as get_lease_for_config."""
    all_rates = _filter_for_state(load_lease_rates(base_dir), state)
    matching = [
        r for r in all_rates
        if r["model_year"] == model_year
        and r["body_style"] == body_style
        and r["tier"] == tier
    ]
    by_trim: dict[str, list[dict]] = {}
    for r in matching:
        by_trim.setdefault(r["trim"], []).append(r)

    result = {}
    for trim, rows in by_trim.items():
        picked = _pick_per_term_lease(rows, state)
        result[trim] = sorted(picked.values(), key=lambda r: r["term"])
    return result
