"""Daily SAP campaign code report import service."""

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.models.transaction import DealTransaction
from app.models.campaign_code import CampaignCode
from app.models.dealer import Dealer


# Common column header mappings
COLUMN_MAP = {
    "campaign code": "campaign_code",
    "campaign_code": "campaign_code",
    "campaign description": "campaign_desc",
    "vin": "vin",
    "material": "material",
    "sales order": "sales_order",
    "sales_order": "sales_order",
    "handover date": "retail_date",
    "retail date": "retail_date",
    "retail_date": "retail_date",
    "ship-to party": "dealer_ship_to",
    "ship_to_party": "dealer_ship_to",
    "ship-to name": "dealer_name",
    "ship_to_name": "dealer_name",
    "channel": "channel",
    "amount": "support_amount",
    "support value": "support_amount",
    "support_amount": "support_amount",
    "trim": "trim",
    "msrp": "msrp",
}


def import_daily_report(db: Session, file_bytes: bytes, filename: str) -> dict:
    """Parse and import a daily SAP campaign code report Excel file."""
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"total_rows": 0, "imported": 0, "duplicates": 0, "errors": 0, "anomalies": 0, "details": []}

    # Map headers
    raw_headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    col_indices = {}
    for idx, header in enumerate(raw_headers):
        mapped = COLUMN_MAP.get(header)
        if mapped:
            col_indices[mapped] = idx

    results = {"total_rows": 0, "imported": 0, "duplicates": 0, "errors": 0, "anomalies": 0, "details": []}

    known_codes = {c.code for c in db.query(CampaignCode.code).all()}
    known_dealers = {d.ship_to_code for d in db.query(Dealer.ship_to_code).all()}

    for row_num, row in enumerate(rows[1:], start=2):
        results["total_rows"] += 1
        try:
            def get_val(field):
                idx = col_indices.get(field)
                if idx is not None and idx < len(row):
                    return row[idx]
                return None

            vin = str(get_val("vin") or "").strip()
            if not vin:
                results["errors"] += 1
                results["details"].append(f"Row {row_num}: Missing VIN")
                continue

            code = str(get_val("campaign_code") or "").strip()
            ship_to = str(get_val("dealer_ship_to") or "").strip()

            # Check for duplicate VIN + code
            existing = db.query(DealTransaction).filter(
                DealTransaction.vin == vin,
                DealTransaction.campaign_code == code,
            ).first()
            if existing:
                results["duplicates"] += 1
                continue

            # Parse retail date
            retail_date = get_val("retail_date")
            if isinstance(retail_date, str):
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
                    try:
                        retail_date = datetime.strptime(retail_date, fmt).date()
                        break
                    except ValueError:
                        continue
            elif hasattr(retail_date, "date"):
                retail_date = retail_date.date()

            # Parse amount
            amount = get_val("support_amount")
            try:
                amount = Decimal(str(amount)) if amount else Decimal("0")
            except Exception:
                amount = Decimal("0")

            # Anomaly detection
            anomalies = []
            if code and code not in known_codes:
                anomalies.append(f"Unknown code: {code}")
            if ship_to and ship_to not in known_dealers:
                anomalies.append(f"Unknown dealer: {ship_to}")

            anomaly_flag = "; ".join(anomalies) if anomalies else None
            if anomalies:
                results["anomalies"] += 1

            txn = DealTransaction(
                vin=vin,
                campaign_code=code,
                dealer_ship_to=ship_to,
                dealer_name=str(get_val("dealer_name") or ""),
                sales_order=str(get_val("sales_order") or ""),
                retail_date=retail_date if retail_date and hasattr(retail_date, "year") else None,
                channel=str(get_val("channel") or ""),
                material=str(get_val("material") or ""),
                trim=str(get_val("trim") or ""),
                msrp=Decimal(str(get_val("msrp"))) if get_val("msrp") else None,
                support_amount=amount,
                source_file=filename,
                anomaly_flag=anomaly_flag,
            )
            db.add(txn)
            results["imported"] += 1

        except Exception as e:
            results["errors"] += 1
            results["details"].append(f"Row {row_num}: {str(e)}")

    db.commit()
    wb.close()
    return results
