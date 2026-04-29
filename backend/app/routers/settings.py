from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database import get_db
from app.models.dealer import Dealer, Product
from app.models.user import User
from app.models.budget import StackingRule
from app.models.vehicle import Vehicle
from app.models.rate_file import SantanderRateFile, SANTANDER_RATE_FILE_KINDS
from app.schemas.dealer import DealerCreate, DealerResponse, ProductCreate, ProductResponse, UserCreate, UserUpdateRequest
from app.auth.security import get_current_user, require_admin, hash_password
from app.services.vehicle_import import import_master_file, import_master_file_from_path
from app.services.stacking import get_stacking_matrix, DEFAULT_STACKING
from app.services.pdf_generator import generate_program_bulletin, generate_quick_reference_card
from fastapi.responses import FileResponse
import os

router = APIRouter(prefix="/api/settings", tags=["settings"])


# --- Dealers ---
@router.get("/dealers")
def list_dealers(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    dealers = db.query(Dealer).order_by(Dealer.region, Dealer.name).all()
    return [DealerResponse.model_validate(d).model_dump() for d in dealers]


@router.post("/dealers")
def create_dealer(req: DealerCreate, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    d = Dealer(**req.model_dump())
    db.add(d)
    db.commit()
    return DealerResponse.model_validate(d)


@router.put("/dealers/{dealer_id}")
def update_dealer(dealer_id: str, req: DealerCreate, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    d = db.query(Dealer).filter(Dealer.id == dealer_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Dealer not found")
    for key, val in req.model_dump().items():
        setattr(d, key, val)
    db.commit()
    return DealerResponse.model_validate(d)


@router.delete("/dealers/{dealer_id}")
def delete_dealer(dealer_id: str, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    d = db.query(Dealer).filter(Dealer.id == dealer_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Dealer not found")
    db.delete(d)
    db.commit()
    return {"message": "Dealer deleted"}


# --- Products ---
@router.get("/products")
def list_products(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    products = db.query(Product).order_by(Product.model_year, Product.body_style, Product.trim).all()
    return [ProductResponse.model_validate(p).model_dump() for p in products]


@router.post("/products")
def create_product(req: ProductCreate, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    p = Product(**req.model_dump())
    db.add(p)
    db.commit()
    return ProductResponse.model_validate(p)


@router.put("/products/{product_id}")
def update_product(product_id: str, req: ProductCreate, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    for key, val in req.model_dump().items():
        setattr(p, key, val)
    db.commit()
    return ProductResponse.model_validate(p)


@router.delete("/products/{product_id}")
def delete_product(product_id: str, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    db.delete(p)
    db.commit()
    return {"message": "Product deleted"}


# --- Users ---
@router.get("/users")
def list_users(db: Session = Depends(get_db), user: User = Depends(require_admin)):
    users = db.query(User).order_by(User.role, User.name).all()
    return [{"id": u.id, "username": u.username, "role": u.role, "name": u.name,
             "region": u.region, "dealer_id": u.dealer_id, "active": u.active} for u in users]


@router.post("/users")
def create_user(req: UserCreate, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    if db.query(User).filter(User.username == req.username).first():
        raise HTTPException(status_code=400, detail="Username already exists")
    u = User(
        username=req.username,
        password_hash=hash_password(req.password),
        role=req.role,
        name=req.name,
        dealer_id=req.dealer_id,
        region=req.region,
    )
    db.add(u)
    db.commit()
    return {"id": u.id, "username": u.username, "role": u.role, "name": u.name}


@router.put("/users/{user_id}")
def update_user(user_id: str, req: UserUpdateRequest, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    data = req.model_dump(exclude_unset=True)
    if "password" in data and data["password"]:
        u.password_hash = hash_password(data.pop("password"))
    else:
        data.pop("password", None)
    for key, val in data.items():
        setattr(u, key, val)
    db.commit()
    return {"id": u.id, "username": u.username, "role": u.role, "name": u.name}


# --- Stacking Rules ---
@router.get("/stacking-rules")
def get_stacking_rules(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return get_stacking_matrix(db)


@router.put("/stacking-rules")
def update_stacking_rules(rules: dict, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    db.query(StackingRule).delete()
    for deal_type, program_types in rules.items():
        all_types = [
            "bonus_cash", "customer_cash", "apr_cash", "lease_cash", "dealer_cash",
            "cvp", "demonstrator", "loyalty", "conquest", "tactical", "other"
        ]
        for pt in all_types:
            db.add(StackingRule(
                deal_type=deal_type, program_type=pt,
                allowed="Y" if pt in program_types else "N"
            ))
    db.commit()
    return get_stacking_matrix(db)


# --- Mileage Degradation (Lease Residual Adjustment) ---
@router.get("/mileage-degradation")
def get_mileage_degradation(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Get mileage-based residual degradation table (points adjustment from 10k base)."""
    rows = db.query(StackingRule).filter(StackingRule.deal_type == "_mileage_degrade").all()
    if rows:
        return {r.program_type: float(r.allowed) for r in rows}
    # Return defaults
    return {"7500": 1, "10000": 0, "12000": -1, "15000": -2}


@router.put("/mileage-degradation")
def update_mileage_degradation(
    table: dict,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    """Update mileage degradation table. Keys are annual miles, values are residual point adjustments."""
    db.query(StackingRule).filter(StackingRule.deal_type == "_mileage_degrade").delete()
    for mileage, points in table.items():
        db.add(StackingRule(
            deal_type="_mileage_degrade",
            program_type=str(mileage),
            allowed=str(points),
        ))
    db.commit()
    return table


# --- Document Generation ---
@router.get("/bulletin/{program_id}")
def get_bulletin(program_id: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    filepath = generate_program_bulletin(db, program_id)
    return FileResponse(filepath, media_type="application/pdf", filename=os.path.basename(filepath))


@router.get("/quick-reference")
def get_quick_reference(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    filepath = generate_quick_reference_card(db)
    return FileResponse(filepath, media_type="application/pdf", filename=os.path.basename(filepath))


@router.get("/public/bulletin/{program_id}")
def get_public_bulletin(program_id: str, db: Session = Depends(get_db)):
    """
    Retailer-facing PDF download. Same generator as the admin route,
    gated to live programs only \u2014 a 404 for drafts and staged programs
    so an unauthenticated visitor can't pull a doc for something that
    isn't customer-facing yet.
    """
    from app.models.program import Program
    prog = db.query(Program).filter(Program.id == program_id).first()
    if not prog or prog.status != "active" or not getattr(prog, "published", False):
        raise HTTPException(status_code=404, detail="Program not found or not yet live")
    filepath = generate_program_bulletin(db, program_id)
    return FileResponse(filepath, media_type="application/pdf", filename=os.path.basename(filepath))


@router.get("/public/quick-reference")
def get_public_quick_reference(db: Session = Depends(get_db)):
    """
    Retailer-facing quick reference card. Generator pulls from active
    programs by default; the gating happens upstream when admins decide
    which programs to publish.
    """
    filepath = generate_quick_reference_card(db)
    return FileResponse(filepath, media_type="application/pdf", filename=os.path.basename(filepath))


# --- Vehicle Inventory (Master File) ---
@router.post("/vehicles/import")
async def import_vehicles(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    """Upload and import vehicle inventory from Master File (.xlsb or .xlsx)."""
    content = await file.read()
    if len(content) > 50 * 1024 * 1024:  # 50MB limit for master files
        raise HTTPException(status_code=400, detail="File too large (max 50MB)")
    result = import_master_file(db, content, file.filename)
    return result


@router.post("/vehicles/import-local")
def import_vehicles_local(db: Session = Depends(get_db), user: User = Depends(require_admin)):
    """Import vehicle inventory from the local Master File in the inputs folder."""
    import glob
    input_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
    patterns = [os.path.join(input_dir, "*.xlsb"), os.path.join(input_dir, "*.xlsx")]
    files = []
    for pat in patterns:
        files.extend(glob.glob(pat))

    master_files = [f for f in files if "master" in os.path.basename(f).lower()]
    if not master_files:
        raise HTTPException(status_code=404, detail=f"No Master File found in {input_dir}")

    result = import_master_file_from_path(db, master_files[0])
    result["file"] = os.path.basename(master_files[0])
    return result


@router.get("/vehicles")
def list_vehicles(
    status: str = Query(None),
    dealer_ship_to: str = Query(None),
    search: str = Query(None),
    skip: int = Query(0),
    limit: int = Query(50),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """List vehicles from inventory with optional filters."""
    q = db.query(Vehicle)
    if status:
        q = q.filter(Vehicle.status == status)
    if dealer_ship_to:
        q = q.filter(Vehicle.dealer_ship_to == dealer_ship_to)
    if search:
        q = q.filter(
            (Vehicle.vin.contains(search.upper())) |
            (Vehicle.dealer_name.ilike(f"%{search}%"))
        )
    total = q.count()
    vehicles = q.order_by(Vehicle.vin).offset(skip).limit(limit).all()
    return {
        "total": total,
        "items": [
            {
                "id": v.id, "vin": v.vin, "model_year": v.model_year,
                "body_style": v.body_style, "trim": v.trim,
                "special_edition": v.special_edition, "msrp": float(v.msrp) if v.msrp else None,
                "dealer_ship_to": v.dealer_ship_to, "dealer_name": v.dealer_name,
                "status": v.status, "color_exterior": v.color_exterior,
            }
            for v in vehicles
        ],
    }


@router.get("/vehicles/stats")
def vehicle_stats(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Get vehicle inventory statistics."""
    total = db.query(Vehicle).count()
    by_status = db.query(Vehicle.status, func.count(Vehicle.id)).group_by(Vehicle.status).all()
    by_model = db.query(Vehicle.model_year, Vehicle.body_style, func.count(Vehicle.id)).group_by(
        Vehicle.model_year, Vehicle.body_style
    ).all()
    return {
        "total": total,
        "by_status": {s: c for s, c in by_status if s},
        "by_model": [{"model_year": m, "body_style": b, "count": c} for m, b, c in by_model if m],
    }


# --- Santander rate file uploads -----------------------------------------
# Admins upload the four monthly Santander input xlsx files via the
# Settings page; the payment calculator reads the latest of each kind
# instead of the filesystem-bundled copies the repo originally
# shipped. One row per kind — uploading replaces the previous file.

_KIND_META = {
    "apr":         {"label": "National APR",    "expected_sheet": "Retail", "filename_must_contain": "APRInput",   "must_not_contain": "State_"},
    "lease":       {"label": "National Lease",  "expected_sheet": "Lease",  "filename_must_contain": "LeaseInput", "must_not_contain": "State_"},
    "state_apr":   {"label": "State APR",       "expected_sheet": "Retail", "filename_must_contain": "APRInput",   "must_contain": "State_"},
    "state_lease": {"label": "State Lease",     "expected_sheet": "Lease",  "filename_must_contain": "LeaseInput", "must_contain": "State_"},
}


def _validate_rate_file(kind: str, filename: str, contents: bytes):
    """Reject obviously-wrong uploads — wrong kind for this slot or a
    file the parser can't open. The filename heuristics catch the
    common 'wrong slot' mistake (uploading the State file under the
    National card or vice versa); the openpyxl probe catches files
    that aren't actually xlsx."""
    meta = _KIND_META[kind]
    name = filename or ""
    if meta["filename_must_contain"] not in name:
        raise HTTPException(
            status_code=400,
            detail=f"Filename must contain '{meta['filename_must_contain']}' for the {meta['label']} slot — got '{name}'."
        )
    if "must_contain" in meta and meta["must_contain"] not in name:
        raise HTTPException(
            status_code=400,
            detail=f"This is the {meta['label']} slot — filename must start with '{meta['must_contain']}' (got '{name}'). Use the National slot instead?"
        )
    if "must_not_contain" in meta and meta["must_not_contain"] in name:
        raise HTTPException(
            status_code=400,
            detail=f"This is the {meta['label']} slot — filename must NOT start with '{meta['must_not_contain']}' (got '{name}'). Use the State slot instead?"
        )
    # Sanity-check by opening the workbook and confirming the expected sheet name.
    import io
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
        if meta["expected_sheet"] not in wb.sheetnames:
            raise HTTPException(
                status_code=400,
                detail=f"Workbook is missing the expected '{meta['expected_sheet']}' sheet (got: {wb.sheetnames})."
            )
        wb.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read xlsx file: {e}")


@router.get("/santander-rates")
def list_santander_rate_files(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """List the four Santander rate-file slots with their currently
    uploaded file's metadata. Empty slots show the kind/label only —
    UI uses this to render upload cards with current-file info."""
    rows = {r.kind: r for r in db.query(SantanderRateFile).all()}
    out = []
    for kind, meta in _KIND_META.items():
        r = rows.get(kind)
        out.append({
            "kind": kind,
            "label": meta["label"],
            "filename": r.filename if r else None,
            "size_bytes": len(r.data) if r and r.data else 0,
            "uploaded_at": r.uploaded_at.isoformat() if r and r.uploaded_at else None,
            "uploaded_by": r.uploaded_by if r else None,
        })
    return out


@router.post("/santander-rates/{kind}")
async def upload_santander_rate_file(
    kind: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    """Upload (or replace) a Santander rate file. Triggers a parse-cache
    invalidation so the next /api/lookup/rates call picks up the new
    rates without a redeploy."""
    if kind not in SANTANDER_RATE_FILE_KINDS:
        raise HTTPException(status_code=400, detail=f"Unknown rate-file kind: {kind}")
    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")
    _validate_rate_file(kind, file.filename or "", contents)

    existing = db.query(SantanderRateFile).filter(SantanderRateFile.kind == kind).first()
    if existing:
        existing.filename = file.filename
        existing.data = contents
        existing.uploaded_by = user.username
    else:
        db.add(SantanderRateFile(
            kind=kind, filename=file.filename, data=contents, uploaded_by=user.username,
        ))
    db.commit()

    # Invalidate the in-process rate cache so subsequent lookups re-parse the new bytes.
    from app.services import santander_rates as sr
    sr.invalidate_cache()

    return {"status": "ok", "kind": kind, "filename": file.filename, "size_bytes": len(contents)}


@router.delete("/santander-rates/{kind}")
def delete_santander_rate_file(
    kind: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    """Clear the upload for a slot. Lookup falls back to the filesystem
    copy bundled in the repo (if any) — useful when a bad upload needs
    rolling back without re-uploading the previous month's file."""
    if kind not in SANTANDER_RATE_FILE_KINDS:
        raise HTTPException(status_code=400, detail=f"Unknown rate-file kind: {kind}")
    row = db.query(SantanderRateFile).filter(SantanderRateFile.kind == kind).first()
    if row:
        db.delete(row)
        db.commit()
        from app.services import santander_rates as sr
        sr.invalidate_cache()
    return {"status": "ok", "kind": kind}
