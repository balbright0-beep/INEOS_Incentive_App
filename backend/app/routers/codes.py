import os
from datetime import datetime, date
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from app.database import get_db
from app.models.campaign_code import CampaignCode, CampaignCodeLayer
from app.models.program import Program
from app.schemas.campaign_code import CampaignCodeResponse, CampaignCodeUpdate
from app.auth.security import get_current_user, require_admin
from app.models.user import User
from app.services.code_matrix import rebuild_matrix
from app.config import settings

# ── INEOS Excel style constants ──
DARK_GREEN = "0A2E1F"
MUSHROOM_ROW_A = "E8DFD0"
MUSHROOM_ROW_B = "F2ECE3"
WARM_TAN = "DDD3C0"
LIGHT_GREEN_TEXT = "7FBF8E"
DARK_TEXT = "333333"
GREEN_TEXT = "0A2E1F"

TITLE_FONT = Font(name="Arial", size=16, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name="Arial", size=9, italic=True, color=LIGHT_GREEN_TEXT)
HEADER_FONT = Font(name="Arial", size=9, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=9, color=DARK_TEXT)
DATA_FONT_BOLD = Font(name="Arial", size=9, bold=True, color=DARK_TEXT)
CODE_FONT = Font(name="Arial", size=9, bold=True, color=GREEN_TEXT)
AMOUNT_FONT = Font(name="Arial", size=10, bold=True, color=GREEN_TEXT)
SECTION_FONT = Font(name="Arial", size=9, bold=True, italic=True, color=GREEN_TEXT)
NOTE_FONT = Font(name="Arial", size=8, italic=True, color="666666")

DARK_GREEN_FILL = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
ROW_FILL_A = PatternFill(start_color=MUSHROOM_ROW_A, end_color=MUSHROOM_ROW_A, fill_type="solid")
ROW_FILL_B = PatternFill(start_color=MUSHROOM_ROW_B, end_color=MUSHROOM_ROW_B, fill_type="solid")
WARM_FILL = PatternFill(start_color=WARM_TAN, end_color=WARM_TAN, fill_type="solid")

THIN_BORDER_BOTTOM = Border(bottom=Side(style="thin", color="C0B8A8"))
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

router = APIRouter(prefix="/api/codes", tags=["codes"])


@router.get("/public")
def list_public_codes(db: Session = Depends(get_db)):
    """
    Public listing of active campaign codes for the retailer finder.
    Filters out any code with at least one staged (unpublished)
    contributing program, mirroring the gate in lookup_incentive so
    customers never see a number admins haven't signed off on.
    """
    codes = db.query(CampaignCode).filter(CampaignCode.active == True).order_by(
        CampaignCode.body_style, CampaignCode.model_year, CampaignCode.deal_type,
        CampaignCode.conquest_flag, CampaignCode.loyalty_flag
    ).all()
    out = []
    for c in codes:
        layers = (
            db.query(CampaignCodeLayer, Program)
            .join(Program, CampaignCodeLayer.program_id == Program.id)
            .filter(CampaignCodeLayer.campaign_code_id == c.id)
            .all()
        )
        if not layers:
            continue
        if not all(getattr(prog, "published", False) for _, prog in layers):
            continue
        out.append({
            "id": c.id,
            "code": c.code,
            "label": c.label,
            "body_style": c.body_style,
            "model_year": c.model_year,
            "trim": c.trim,
            "deal_type": c.deal_type,
            "loyalty_flag": bool(c.loyalty_flag),
            "conquest_flag": bool(c.conquest_flag),
            "special_flag": c.special_flag,
            "support_amount": float(c.support_amount or 0),
            "effective_date": c.effective_date.isoformat() if c.effective_date else None,
            "expiration_date": c.expiration_date.isoformat() if c.expiration_date else None,
            "layers": [
                {
                    "program_name": prog.name,
                    "program_type": prog.program_type,
                    "layer_amount": float(layer.layer_amount),
                }
                for layer, prog in layers
            ],
        })
    return out


@router.get("")
def list_codes(
    model_year: str = Query(None),
    body_style: str = Query(None),
    deal_type: str = Query(None),
    active: bool = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = db.query(CampaignCode)
    if model_year:
        q = q.filter(CampaignCode.model_year == model_year)
    if body_style:
        q = q.filter(CampaignCode.body_style == body_style)
    if deal_type:
        q = q.filter(CampaignCode.deal_type == deal_type)
    if active is not None:
        q = q.filter(CampaignCode.active == active)
    codes = q.order_by(CampaignCode.code).all()

    result = []
    for c in codes:
        layers = (
            db.query(CampaignCodeLayer, Program.name)
            .join(Program, CampaignCodeLayer.program_id == Program.id)
            .filter(CampaignCodeLayer.campaign_code_id == c.id)
            .all()
        )
        result.append({
            **CampaignCodeResponse.model_validate(c).model_dump(),
            "layers": [
                {"id": l.CampaignCodeLayer.id, "program_id": l.CampaignCodeLayer.program_id,
                 "program_name": l.name, "layer_amount": float(l.CampaignCodeLayer.layer_amount)}
                for l in layers
            ],
        })
    return result


@router.post("/rebuild")
def rebuild_codes(db: Session = Depends(get_db), user: User = Depends(require_admin)):
    matrix = rebuild_matrix(db, preview_only=False)
    return {"message": "Matrix rebuilt", "total_codes": len(matrix)}


@router.get("/preview")
def preview_codes(db: Session = Depends(get_db), user: User = Depends(require_admin)):
    matrix = rebuild_matrix(db, preview_only=True)
    return matrix


@router.put("/{code_id}")
def update_code(
    code_id: str, req: CampaignCodeUpdate,
    db: Session = Depends(get_db), user: User = Depends(require_admin),
):
    code = db.query(CampaignCode).filter(CampaignCode.id == code_id).first()
    if not code:
        raise HTTPException(status_code=404, detail="Code not found")
    for key, val in req.model_dump(exclude_unset=True).items():
        setattr(code, key, val)
    db.commit()
    return CampaignCodeResponse.model_validate(code)


def _build_headline(code, layers_data):
    """Build the HEADLINE string like '$5,000 Consumer Cash + $1,500 Loyalty Bonus'."""
    if not layers_data:
        amount = float(code.support_amount or 0)
        if amount == 0:
            dt = (code.deal_type or "").lower()
            if dt == "apr":
                return "Santander Subvented APR"
            elif dt == "lease":
                return "Santander Subvented Lease"
            elif dt == "demo":
                return "Demonstrator"
            return "No Sales Campaign Applied"
        return f"${amount:,.0f}"

    parts = []
    for layer, prog in layers_data:
        amount = float(layer.layer_amount)
        ptype = prog.program_type.replace("_", " ").title()
        if amount > 0:
            parts.append(f"${amount:,.0f} {ptype}")
        else:
            parts.append(ptype)
    return " + ".join(parts) if parts else ""


def _campaign_label(code, layers_data):
    """Extract campaign layer names for CAMPAIGN 1-4 columns."""
    labels = []
    for layer, prog in layers_data:
        labels.append(prog.name.replace("April 2026 ", "").replace("2026 ", ""))
    return labels


def _model_code(body_style, model_year):
    """Generate model code like G01C (SW MY25), G09C (QM MY25), G01D (SW MY26)."""
    body = "G01" if body_style == "station_wagon" else "G09"
    suffix = "D" if model_year and model_year >= "MY26" else "C"
    return f"{body}{suffix}"


def _customer_group(deal_type):
    """Map deal type to customer group."""
    mapping = {
        "cash": "2 Private Retailer",
        "apr": "2 Private Retailer",
        "lease": "9 Leasing",
        "cvp": "Courtesy Car",
        "demo": "5 Demo",
    }
    return mapping.get(deal_type, "2 Private Retailer")


def _vehicle_label(code):
    """Build the vehicle description like '2025 Station Wagon' or 'APR Program (2026 Station Wagon)'."""
    my = (code.model_year or "MY25").replace("MY", "20")
    body = (code.body_style or "station_wagon").replace("_", " ").title()
    dt = (code.deal_type or "cash").lower()
    special = code.special_flag

    if special:
        return f"{my} {special.replace('_', ' ').title()}"
    if dt == "apr":
        return f"APR Program ({my} {body})"
    elif dt == "lease":
        return f"Lease Program ({my} {body})"
    elif dt == "cvp":
        return "Courtesy Vehicle Program"
    elif dt == "demo":
        return "Demonstrator"
    return f"{my} {body}"


@router.get("/export")
def export_codes(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return _build_codes_xlsx(db, published_only=False)


@router.get("/public/export")
def export_public_codes(db: Session = Depends(get_db)):
    """
    Excel extract for the retailer-facing finder. Same shape as the
    admin export; codes whose contributing programs are all live
    (published=True) only \u2014 staged programs are filtered out so the
    download always matches what the public lookup is showing.
    """
    return _build_codes_xlsx(db, published_only=True)


def _build_codes_xlsx(db: Session, published_only: bool = False):
    codes = db.query(CampaignCode).filter(CampaignCode.active == True).order_by(
        CampaignCode.body_style, CampaignCode.model_year, CampaignCode.deal_type,
        CampaignCode.conquest_flag, CampaignCode.loyalty_flag
    ).all()

    # Pre-fetch all layer data
    code_layers = {}
    for c in codes:
        layers = (
            db.query(CampaignCodeLayer, Program)
            .join(Program, CampaignCodeLayer.program_id == Program.id)
            .filter(CampaignCodeLayer.campaign_code_id == c.id)
            .all()
        )
        code_layers[c.id] = layers

    # Public extract gates on every contributing program being published.
    # Codes with any staged layer are dropped entirely; we don't render a
    # partial breakdown because the totals would be misleading.
    if published_only:
        codes = [
            c for c in codes
            if code_layers.get(c.id) and all(getattr(prog, "published", False) for _, prog in code_layers[c.id])
        ]

    # Get date range from active programs
    active_programs = db.query(Program).filter(Program.status == "active").all()
    eff_date = min((p.effective_date for p in active_programs), default=date.today())
    exp_date = max((p.expiration_date for p in active_programs), default=date.today())
    month_year = eff_date.strftime("%B %Y")

    wb = Workbook()

    # ════════════ Sheet 1: USA Campaign Matrix ════════════
    ws = wb.active
    ws.title = "USA Campaign Matrix"
    ws.sheet_properties.tabColor = DARK_GREEN

    # Column widths matching INEOS format
    col_widths = {
        "A": 28, "B": 22, "C": 22, "D": 16,   # Campaign 1-4
        "E": 45,                                   # Headline
        "F": 18,                                   # Customer Group
        "G": 14,                                   # Sales Code
        "H": 12,                                   # Model Year
        "I": 12,                                   # Model Codes
        "J": 16,                                   # Campaign Support
        "K": 12,                                   # Fixed Margin
        "L": 14,                                   # Variable Margin
        "M": 20,                                   # Notes
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    NUM_COLS = 13

    # ── Row 1: Title (merged, dark green) ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    title_cell = ws.cell(row=1, column=1,
                         value=f"{month_year} Campaign Matrix - United States of America")
    title_cell.font = TITLE_FONT
    title_cell.fill = DARK_GREEN_FILL
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32
    for col in range(2, NUM_COLS + 1):
        ws.cell(row=1, column=col).fill = DARK_GREEN_FILL

    # ── Row 2: Subtitle (merged, dark green, italic) ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
    def _ordinal(d):
        day = d.day
        suffix = "th" if 11 <= day <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
        return f"{day}{suffix} {d.strftime('%B %Y')}"
    sub_cell = ws.cell(row=2, column=1,
                       value=f" (Valid for retail sales taken {_ordinal(eff_date)} to {_ordinal(exp_date)} & Handover by {exp_date.strftime('%B %d, %Y')})")
    sub_cell.font = SUBTITLE_FONT
    sub_cell.fill = DARK_GREEN_FILL
    for col in range(2, NUM_COLS + 1):
        ws.cell(row=2, column=col).fill = DARK_GREEN_FILL

    # ── Row 3: Blank spacer ──
    ws.row_dimensions[3].height = 8

    # ── Row 4: Column Headers ──
    headers = ["CAMPAIGN 1", "CAMPAIGN 2", "CAMPAIGN 3", "CAMPAIGN 4",
               "HEADLINE", "CUSTOMER GROUP", "SALES CODE",
               "Eligible Model Year", "Eligible Model Codes",
               "Campaign Support", "Fixed Margin", "Variable Margin", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = DARK_GREEN_FILL
        cell.alignment = HEADER_ALIGNMENT
    ws.row_dimensions[4].height = 28

    # ── Row 5: Sub-header row (payment method labels) ──
    sub_labels = {7: "Paid Via", 10: "Credit Note", 11: "Invoice", 12: "Credit Note"}
    for col, label in sub_labels.items():
        cell = ws.cell(row=5, column=col, value=label)
        cell.font = SECTION_FONT
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows starting at row 6 ──
    current_row = 6
    last_vehicle = None

    for c in codes:
        layers = code_layers.get(c.id, [])
        vehicle = _vehicle_label(c)
        headline = _build_headline(c, layers)
        campaign_labels = _campaign_label(c, layers)
        model_code = _model_code(c.body_style, c.model_year)
        customer_group = _customer_group(c.deal_type)
        amount = float(c.support_amount or 0)

        # Insert section separator when vehicle group changes
        if last_vehicle and vehicle.split("(")[0].strip() != last_vehicle.split("(")[0].strip():
            # Check if deal type category changed
            last_dt = ""
            if "APR" in str(last_vehicle):
                last_dt = "apr"
            elif "Lease" in str(last_vehicle):
                last_dt = "lease"
            curr_dt = ""
            if "APR" in vehicle:
                curr_dt = "apr"
            elif "Lease" in vehicle:
                curr_dt = "lease"
            if last_dt != curr_dt or ("Courtesy" in vehicle) or ("Demonstrator" in vehicle):
                current_row += 1  # blank separator row

        # Alternate row colors
        row_fill = ROW_FILL_A if (current_row % 2 == 0) else ROW_FILL_B

        # Campaign 1 (vehicle label)
        ws.cell(row=current_row, column=1, value=vehicle).font = DATA_FONT_BOLD

        # Campaign 2-4 (layer names)
        for idx, label in enumerate(campaign_labels[:3]):
            ws.cell(row=current_row, column=2 + idx, value=label).font = DATA_FONT

        # Headline
        ws.cell(row=current_row, column=5, value=headline).font = DATA_FONT

        # Customer Group
        ws.cell(row=current_row, column=6, value=customer_group).font = DATA_FONT

        # Sales Code (THE CODE - bold green)
        ws.cell(row=current_row, column=7, value=c.code).font = CODE_FONT

        # Model Year
        ws.cell(row=current_row, column=8, value=c.model_year or "").font = DATA_FONT

        # Model Codes
        ws.cell(row=current_row, column=9, value=model_code).font = DATA_FONT

        # Campaign Support (amount - bold green)
        if amount > 0:
            amt_cell = ws.cell(row=current_row, column=10, value=amount)
            amt_cell.font = AMOUNT_FONT
            amt_cell.number_format = '#,##0'
        else:
            ws.cell(row=current_row, column=10, value="").font = DATA_FONT

        # Fixed Margin
        body = (c.body_style or "").lower()
        margin = 0.08 if body == "station_wagon" else 0.06
        margin_cell = ws.cell(row=current_row, column=11, value=margin)
        margin_cell.font = DATA_FONT
        margin_cell.number_format = '0.00'

        # Variable Margin (blank for most)
        ws.cell(row=current_row, column=12, value="").font = DATA_FONT

        # Notes
        ws.cell(row=current_row, column=13, value="").font = DATA_FONT

        # Apply row fill and border
        for col in range(1, NUM_COLS + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = row_fill
            cell.border = THIN_BORDER_BOTTOM
            if not cell.alignment or cell.alignment.horizontal is None:
                cell.alignment = Alignment(vertical="center")

        # Center-align certain columns
        for col in [7, 8, 9, 10, 11, 12]:
            ws.cell(row=current_row, column=col).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        last_vehicle = vehicle
        current_row += 1

    # ── Footer row ──
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=NUM_COLS)
    footer = ws.cell(row=current_row, column=1,
                     value=f"Confidential - INEOS Automotive Americas - Generated {datetime.now().strftime('%B %d, %Y')}")
    footer.font = NOTE_FONT
    footer.alignment = Alignment(horizontal="center")

    # ── Freeze panes (freeze headers) ──
    ws.freeze_panes = "A6"

    # ── Print settings ──
    ws.sheet_properties.pageSetUpPr = None
    ws.print_title_rows = "1:5"

    # ════════════ Sheet 2: Code Lookup (flat reference) ════════════
    ws2 = wb.create_sheet("Code Lookup")
    ws2.sheet_properties.tabColor = "4A8C42"

    lookup_headers = ["SALES CODE", "VEHICLE", "DEAL TYPE", "LOYALTY", "CONQUEST",
                      "CAMPAIGN SUPPORT", "HEADLINE", "EFFECTIVE", "EXPIRATION"]
    for col, h in enumerate(lookup_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = DARK_GREEN_FILL
        cell.alignment = HEADER_ALIGNMENT

    for row_num, c in enumerate(codes, 2):
        layers = code_layers.get(c.id, [])
        ws2.cell(row=row_num, column=1, value=c.code).font = CODE_FONT
        ws2.cell(row=row_num, column=2, value=_vehicle_label(c)).font = DATA_FONT
        ws2.cell(row=row_num, column=3, value=(c.deal_type or "").upper()).font = DATA_FONT
        ws2.cell(row=row_num, column=4, value="Yes" if c.loyalty_flag else "").font = DATA_FONT
        ws2.cell(row=row_num, column=5, value="Yes" if c.conquest_flag else "").font = DATA_FONT
        amt = float(c.support_amount or 0)
        amt_cell = ws2.cell(row=row_num, column=6, value=amt if amt > 0 else "")
        amt_cell.font = AMOUNT_FONT if amt > 0 else DATA_FONT
        amt_cell.number_format = '$#,##0' if amt > 0 else ''
        ws2.cell(row=row_num, column=7, value=_build_headline(c, layers)).font = DATA_FONT
        ws2.cell(row=row_num, column=8, value=str(c.effective_date) if c.effective_date else "").font = DATA_FONT
        ws2.cell(row=row_num, column=9, value=str(c.expiration_date) if c.expiration_date else "").font = DATA_FONT

        fill = ROW_FILL_A if row_num % 2 == 0 else ROW_FILL_B
        for col in range(1, 10):
            ws2.cell(row=row_num, column=col).fill = fill
            ws2.cell(row=row_num, column=col).border = THIN_BORDER_BOTTOM
            ws2.cell(row=row_num, column=col).alignment = Alignment(vertical="center")

    lookup_widths = [14, 32, 10, 8, 10, 16, 48, 12, 12]
    for i, w in enumerate(lookup_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ════════════ Sheet 3: Layer Breakdown ════════════
    ws3 = wb.create_sheet("Layer Breakdown")
    ws3.sheet_properties.tabColor = "D9D7D0"

    layer_headers = ["SALES CODE", "SUPPORT TOTAL", "LAYER", "PROGRAM", "LAYER AMOUNT"]
    for col, h in enumerate(layer_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = DARK_GREEN_FILL
        cell.alignment = HEADER_ALIGNMENT

    lr = 2
    for c in codes:
        layers = code_layers.get(c.id, [])
        if not layers:
            continue
        for idx, (layer, prog) in enumerate(layers):
            ws3.cell(row=lr, column=1, value=c.code if idx == 0 else "").font = CODE_FONT if idx == 0 else DATA_FONT
            ws3.cell(row=lr, column=2, value=float(c.support_amount) if idx == 0 else "").font = AMOUNT_FONT if idx == 0 else DATA_FONT
            if idx == 0:
                ws3.cell(row=lr, column=2).number_format = '$#,##0'
            ws3.cell(row=lr, column=3, value=idx + 1).font = DATA_FONT
            ws3.cell(row=lr, column=4, value=prog.name).font = DATA_FONT
            amt_cell = ws3.cell(row=lr, column=5, value=float(layer.layer_amount))
            amt_cell.font = DATA_FONT
            amt_cell.number_format = '$#,##0'
            fill = ROW_FILL_A if lr % 2 == 0 else ROW_FILL_B
            for col in range(1, 6):
                ws3.cell(row=lr, column=col).fill = fill
                ws3.cell(row=lr, column=col).border = THIN_BORDER_BOTTOM
            lr += 1

    layer_widths = [14, 16, 8, 35, 14]
    for i, w in enumerate(layer_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    # ════════════ Save & return ════════════
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Campaign_Code_Matrix_{month_year.replace(' ', '_')}_V1.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
